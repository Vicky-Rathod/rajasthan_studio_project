from django.contrib.auth.models import User
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
import os
from django.conf import settings
from django.core.files.storage import default_storage
import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
import datetime
from .models import Appointment
from .serializers import AppointmentSerializer
from rest_framework import  viewsets

def get_json():

	# Gets the Dataset.xlsx from Media directory root of  the project 
	dataset = open(os.path.join(settings.MEDIA_ROOT, 'Dataset.xlsx'), 'rb')
	wb = load_workbook(dataset)

	# Gets the Dataset worksheet

	sheet = wb['Dataset']

	# List to hold dictionaries
	coach_list = []
	def default(o):
		if isinstance(o, (datetime.date, datetime.datetime)):
			return o.isoformat()


	# Iterate through each row in worksheet and fetch values into dict
	for row in islice(sheet.values, 1, sheet.max_row):
		coach = OrderedDict()
		coach['name'] = row[0]

		coach['timezone'] = row[1]
		coach['day_of_week'] = row[2]
		coach['available_at'] = str(row[3])
		coach['available_until'] = str(row[4])
		coach_list.append(coach)



	j = json.dumps(coach_list,  sort_keys=True,
	indent=1,
	default=default)
	jsons = json.loads(j)
	return jsons


# returns the list of 30 mins time slots by taking Available at , Available Until and Time Delta argument  

def datetime_range(start, end, delta):
	current = start
	while current < end:
		yield current
		current += delta





class CoachView(APIView):


	permission_classes = [AllowAny]

	def get(self, request, format=None):
		"""
		Return a list of all Coached
		"""
		coach_list = get_json()

		return Response(coach_list)


class ThirtyMinView(APIView):


	permission_classes = [AllowAny]

	def post(self, request, format=None):
		"""
		Return a list of all coaches with with there 30 min time schedule.
		"""
		dic = {}
		try :
			coach_list = get_json()
			coach = request.data["name"]
			day = request.data.get("day")
			#time = request.data.get("time","None")

			for i in range(len(coach_list)):
				if coach == coach_list[i]["name"] and day == coach_list[i]["day_of_week"]:
					dts = [dt.strftime('%H:%M') for dt in
		datetime_range(datetime.datetime.strptime(coach_list[i]["available_at"], '%H:%M:%S'), datetime.datetime.strptime(coach_list[i]["available_until"], '%H:%M:%S'),
		datetime.timedelta(minutes=30))]
					dic['name'] = coach_list[i]["name"]
					dic['timezone'] = coach_list[i]["timezone"]
					dic['day_of_week'] = coach_list[i]["day_of_week"]
					res =  [i +" - " + j for i, j in zip(dts[::1], dts[1::1])]
					dic['time_slot'] =  res
		except:
			dic["Error"] = "Please Enter Valid Name and Day"
		return Response([dic])


class AppointmentView(APIView):
	def get(self, request):
		appointment = Appointment.objects.all()
		# the many param informs the serializer that it will be serializing more than a single article.
		serializer = AppointmentSerializer(appointment, many=True)
		return Response({"Appointments": serializer.data})

	def post(self , request):
		user = User.objects.get(id = 1 )
		try:
			serializer = AppointmentSerializer(data = request.data)
			if not serializer.is_valid():
				return Response({
					'status' : 403,
					'errors' : serializer.errors
				})
			serializer.save(user = user)
			return Response({'status' : 200 ,'message' : 'Appointment created successfully'})
		except Exception as e:
			print(e)
			return Response({'status' : 404,'error' : 'something went wrong'})